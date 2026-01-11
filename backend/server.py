from fastapi import FastAPI, APIRouter, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from pydantic import BaseModel
import uuid
from datetime import datetime, timezone
import hashlib
import json
import re

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Excel storage is optional. Set USE_EXCEL=true locally to enable Excel files.
USE_EXCEL = os.environ.get('USE_EXCEL', 'false').lower() == 'true'

app = FastAPI()
api_router = APIRouter(prefix="/api")

EXCEL_DIR = ROOT_DIR / 'data'
EXCEL_DIR.mkdir(exist_ok=True)
STUDENTS_FILE = EXCEL_DIR / 'student_registrations.xlsx'
VOLUNTEERS_FILE = EXCEL_DIR / 'volunteer_registrations.xlsx'
CONTACTS_FILE = EXCEL_DIR / 'contact_messages.xlsx'
COUNTS_FILE = EXCEL_DIR / 'counts.json'
MAX_REGISTRATIONS = 1000
ADMIN_PASSWORD_HASH = hashlib.sha256("admin123".encode()).hexdigest()

if USE_EXCEL:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        raise ImportError("openpyxl is required when USE_EXCEL=true; install it in backend/requirements.txt")

    def init_excel_file(filepath, headers):
        if not filepath.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            wb.save(filepath)

    init_excel_file(STUDENTS_FILE, ['ID', 'Name', 'Age', 'School', 'Email', 'Phone', 'Timestamp'])
    init_excel_file(VOLUNTEERS_FILE, ['ID', 'Name', 'Email', 'Phone', 'School/Organization', 'Timestamp'])
    init_excel_file(CONTACTS_FILE, ['ID', 'Name', 'Email', 'Message', 'Timestamp'])
    init_excel_file(EVENT_FILE if (EVENT_FILE := EXCEL_DIR / 'event_content.xlsx') else EVENT_FILE, ['title', 'description', 'date', 'location'])
else:
    # Ensure counts.json exists for Netlify/production mode (no Excel persistence)
    if not COUNTS_FILE.exists():
        with open(COUNTS_FILE, 'w') as f:
            json.dump({'students': 0, 'volunteers': 0}, f)

class StudentRegistration(BaseModel):
    name: str
    age: str
    school: str
    email: str
    phone: str
    consent: bool

class VolunteerRegistration(BaseModel):
    name: str
    email: str
    phone: str
    organization: str

class ContactMessage(BaseModel):
    name: str
    email: str
    message: str

class EventContent(BaseModel):
    title: str
    description: str
    date: str
    location: str

class AdminLogin(BaseModel):
    password: str

class AdminToken(BaseModel):
    token: str

class RegistrationCount(BaseModel):
    students: int
    volunteers: int
    students_limit_reached: bool
    volunteers_limit_reached: bool

# Helpers for JSON-backed counts (used when USE_EXCEL is false)
def _ensure_counts():
    if not COUNTS_FILE.exists():
        counts = {'students': 0, 'volunteers': 0}
        with open(COUNTS_FILE, 'w') as f:
            json.dump(counts, f)
        return counts
    with open(COUNTS_FILE, 'r') as f:
        return json.load(f)

def _save_counts(counts):
    with open(COUNTS_FILE, 'w') as f:
        json.dump(counts, f)

def get_row_count(filepath):
    if not USE_EXCEL:
        counts = _ensure_counts()
        return counts.get('students', 0) if filepath == STUDENTS_FILE else counts.get('volunteers', 0)

    if not filepath.exists():
        return 0
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    count = ws.max_row - 1
    wb.close()
    return max(0, count)

def add_to_excel(filepath, data):
    # When Excel is disabled, update the simple JSON counts and log the submission
    if not USE_EXCEL:
        counts = _ensure_counts()
        if filepath == STUDENTS_FILE:
            counts['students'] = counts.get('students', 0) + 1
        elif filepath == VOLUNTEERS_FILE:
            counts['volunteers'] = counts.get('volunteers', 0) + 1
        _save_counts(counts)
        logging.info('Submission received (Excel disabled): %s', data)
        return

    wb = load_workbook(filepath)
    ws = wb.active
    ws.append(data)
    wb.save(filepath)
    wb.close()


def read_event_file(filepath):
    """Return event dict or None if not set"""
    if not filepath.exists():
        return None
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    if ws.max_row < 2:
        wb.close()
        return None
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    values = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
    wb.close()
    return dict(zip(headers, values))


def write_event_file(filepath, event_dict):
    wb = Workbook()
    ws = wb.active
    ws.append(['title', 'description', 'date', 'location'])
    ws.append([event_dict.get('title'), event_dict.get('description'), event_dict.get('date'), event_dict.get('location')])
    wb.save(filepath)
    wb.close()

@api_router.get("/")
async def root():
    return {"message": "Non-Profit Competition Portal API"}

@api_router.get("/registrations/count", response_model=RegistrationCount)
async def get_registration_count():
    students_count = get_row_count(STUDENTS_FILE)
    volunteers_count = get_row_count(VOLUNTEERS_FILE)
    return RegistrationCount(
        students=students_count,
        volunteers=volunteers_count,
        students_limit_reached=students_count >= MAX_REGISTRATIONS,
        volunteers_limit_reached=volunteers_count >= MAX_REGISTRATIONS
    )

@api_router.post("/students/register")
async def register_student(student: StudentRegistration):
    students_count = get_row_count(STUDENTS_FILE)
    if students_count >= MAX_REGISTRATIONS:
        raise HTTPException(status_code=400, detail="Registration limit reached")

    if not student.consent:
        raise HTTPException(status_code=400, detail="Consent is required")

    # Phone validation: strip non-digits and enforce max 10 digits
    phone_digits = re.sub(r'\D', '', student.phone or '')
    if len(phone_digits) > 10:
        raise HTTPException(status_code=400, detail="Phone number must be at most 10 digits")

    student_id = str(uuid.uuid4())[:8]
    timestamp = datetime.now(timezone.utc).isoformat()

    add_to_excel(STUDENTS_FILE, [
        student_id,
        student.name,
        student.age,
        student.school,
        student.email,
        phone_digits,
        timestamp
    ])

    return {"message": "Registration successful", "id": student_id}

@api_router.post("/volunteers/register")
async def register_volunteer(volunteer: VolunteerRegistration):
    volunteers_count = get_row_count(VOLUNTEERS_FILE)
    if volunteers_count >= MAX_REGISTRATIONS:
        raise HTTPException(status_code=400, detail="Registration limit reached")

    # Phone validation: strip non-digits and enforce max 10 digits
    phone_digits = re.sub(r'\D', '', volunteer.phone or '')
    if len(phone_digits) > 10:
        raise HTTPException(status_code=400, detail="Phone number must be at most 10 digits")

    volunteer_id = str(uuid.uuid4())[:8]
    timestamp = datetime.now(timezone.utc).isoformat()

    add_to_excel(VOLUNTEERS_FILE, [
        volunteer_id,
        volunteer.name,
        volunteer.email,
        phone_digits,
        volunteer.organization,
        timestamp
    ])

    return {"message": "Registration successful", "id": volunteer_id}

@api_router.post("/contact")
async def submit_contact(contact: ContactMessage):
    contact_id = str(uuid.uuid4())[:8]
    timestamp = datetime.now(timezone.utc).isoformat()
    
    add_to_excel(CONTACTS_FILE, [
        contact_id,
        contact.name,
        contact.email,
        contact.message,
        timestamp
    ])
    
    return {"message": "Message sent successfully", "id": contact_id}

@api_router.post("/admin/login", response_model=AdminToken)
async def admin_login(credentials: AdminLogin):
    password_hash = hashlib.sha256(credentials.password.encode()).hexdigest()
    if password_hash != ADMIN_PASSWORD_HASH:
        raise HTTPException(status_code=401, detail="Invalid password")
    
    token = hashlib.sha256(f"{credentials.password}{datetime.now(timezone.utc)}".encode()).hexdigest()
    return AdminToken(token=token)

@api_router.get("/admin/event", response_model=EventContent)
async def get_event_content():
    event = read_event_file(EVENT_FILE)
    if not event:
        default_event = {
            "title": "Annual Spell-Bee Competition 2025",
            "description": "Join us for an exciting spelling competition showcasing English language proficiency. Open to students from grades 3-12.",
            "date": "March 15, 2025",
            "location": "Community Center Auditorium"
        }
        write_event_file(EVENT_FILE, default_event)
        return EventContent(**default_event)
    return EventContent(**event) 

@api_router.put("/admin/event")
async def update_event_content(event: EventContent):
    write_event_file(EVENT_FILE, event.model_dump())
    return {"message": "Event updated successfully"} 

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
